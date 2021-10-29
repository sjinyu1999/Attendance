import openpyxl
import numpy as np
import re
'''import os
import win32com.client as win32
print(os.getcwd())
path=os.getcwd()
path1="\\09月汇总表"
fname = path+path1
print(fname)
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
#FileFormat = 51 is for .xlsx extension
wb.SaveAs(fname, FileFormat = 51)
#FileFormat = 56 is for .xls extension
#wb.SaveAs(fname[:-1], FileFormat = 56)
wb.Close()
excel.Application.Quit()
print("转换成功！！！！")'''
#删除原格式的表“09月汇总表.xls”
#os.remove("09月汇总表.xls")
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename='09月汇总表.xlsx')
#print(wb.sheetnames)
ws=wb['刷卡记录']
#print(ws)
print('记录格式：姓名，日期，上班时间，下班时间，备注，工号，部门')
print('[格式："戚厚之", "2021-09-07", "09:27", "18:54", "", "64", ""]')
minrow=ws.min_row #最小行
maxrow=ws.max_row #最大行
mincol=ws.min_column #最小列
maxcol=ws.max_column #最大列
m=3
n=4
for i in range(mincol,maxcol+1):
    #录入姓名
    m+=2
    name='"'+str(ws.cell(row=m,column=11).value)+'"'
    #print(c1)
    #录入日期
    date=ws.cell(row=4,column=i).value
    #数值保留两位
    date="{0:02d}".format(date)
    date='"'+"2021-09-"+str(date)+'"'
    #录入工号
    number='"'+str(ws.cell(row=m,column=3).value)+'"'
    #录入部门
    apartment='"'+str(ws.cell(row=m,column=20).value)+'"'
    #录入上下班时间
    n+=2
    for j in range(mincol-1,maxcol):
        j=j+1
        c3=str(ws.cell(row=n,column=j).value)
        #print(c3)
        #利用正则表达式拆分单元格的数据
        c31=c3[0:5]
        c32=c3[6:11]
        c33=c3[12:17]
        #print(c31,c32,c33,c34)
        #print(c31[0:2])
        goin=[]
        goout=[]
        if c31[0:2]=="09":
           #print("上班时间:"+c31)
           goin="上班时间:"+'"'+c31+'"'
            #print(c31)
        elif c31[0:2]=="18":
           #print("下班时间:"+c31)
           goout="下班时间:"+'"'+c31+'"'
        elif c31[0:2]=="19":
           #print("下班时间:"+c31)
           goout="下班时间:"+'"'+c31+'"'
        elif c31[0:2]=="20":
           #print("下班时间:"+c31)
           goout="下班时间:"+'"'+c31+'"'
        elif c31[0:2]=="21":
           #print("下班时间:"+c31)
           goout="下班时间:"+'"'+c31+'"'
        else:
            pass
        if c32[0:2]=="09":
           #print("上班时间:"+c32)
           goin="上班时间:"+'"'+c32+'"'
        elif c32[0:2]=="18":
           #print("下班时间:"+c32)
           goout="下班时间:"+'"'+c32+'"'
        elif c32[0:2]=="19":
           #print("下班时间:"+c32)
           goout="下班时间:"+'"'+c32+'"'
        elif c32[0:2]=="20":
           #print("下班时间:"+c32)
           goout="下班时间:"+'"'+c32+'"'
        elif c32[0:2]=="21":
           #print("下班时间:"+c32)
           goout="下班时间:"+'"'+c32+'"'
        else:
            pass
        if c33[0:2]=="09":
           #print("上班时间:"+c33)
           goin="上班时间:"+'"'+c33+'"'
        elif c33[0:2]=="18":
           #print("下班时间:"+c33)
           goout="下班时间:"+'"'+c33+'"'
        elif c33[0:2]=="19":
           #print("下班时间:"+c33)
           goout="下班时间:"+'"'+c33+'"'
        elif c33[0:2]=="20":
           #print("下班时间:"+c33)
           goout="下班时间:"+'"'+c33+'"'
        elif c33[0:2]=="21":
           #print("下班时间:"+c33)
           goout="下班时间:"+'"'+c33+'"'
        else:
            pass
        goin1=''.join(goin)
        #print(goin1)
        goout1=''.join(goout)
        #print(goout1)
        if goin1==None:
            #print("")
            goin1='""'
        else:
            goin1=goin1       
        if goout1==None:
            #print("")
            goout1='""'
        else:
            goout1=goout1
        print("姓名:"+name+",日期:"+date+",上班时间:"+goin1+",下班时间:"+goout1+',备注:""'+",工号:"+number+",部门:"+apartment)
        #print("姓名:"+name+",日期:"+date+',备注:""'+",工号:"+number+",部门:"+apartment)


