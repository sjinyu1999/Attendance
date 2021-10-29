import openpyxl
import random
#生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
wb = openpyxl.Workbook()
# 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
ws = wb.active
#更改工作表ws的title
ws.title = 'output'
#对ws的单个单元格传入数据
ws['B2'] = random.random()
from openpyxl.styles import Alignment
#horizontal代表水平方向,vertical代表垂直方向
align = Alignment(horizontal='center',vertical='center')
ws['B2'].alignment = align
from openpyxl.styles import Font
ws['B2'].font = Font(bold=True,color="FF0000")
#ws.append([1,2,3])
# 获取名为'output'的sheet页
sheet1=wb['output']
#写入数据
j=2
print('打印结果:')
for i in range(0,100):
    j+=i
    m=j+i
    #print(j,m)
    #合并单元格
    #print("正在合并单元格。。。")
    ws.merge_cells(start_row=j,start_column=2,end_row=m,end_column=2)
    ws.merge_cells(start_row=2,start_column=j,end_row=2,end_column=m)
    ws.cell(2,j).alignment = align
    ws.cell(j,2).alignment = align
    sheet1.cell(row=2,column=j,value=random.random())
    sheet1.cell(row=j,column=2,value=random.random())
    print("列："+str(sheet1.cell(2,j).value))
    print("行："+str(sheet1.cell(j,2).value))
wb.save('output.xlsx')