import os
'''import win32com.client as win32
#print(os.getcwd())
path=os.getcwd()
path1="\\09月汇总表"
fname = path+path1
#print(fname)
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
# FileFormat = 51 is for .xlsx extension
wb.SaveAs(fname, FileFormat = 51)
# FileFormat = 56 is for .xls extension
# wb.SaveAs(fname[:-1], FileFormat = 56)
wb.Close()
excel.Application.Quit()
print("转换成功！！！！")'''
# 删除原格式的表“09月汇总表.xls”
# os.remove("09月汇总表.xls")
from openpyxl import load_workbook
wb = load_workbook(filename='09月汇总表.xlsx')
# print(wb.sheetnames)
ws = wb['刷卡记录']
# print(ws)
print('记录格式：姓名，日期，上班时间，下班时间，备注，工号，部门')
print('[格式："戚厚之", "2021-09-07", "09:27", "18:54", "", "64", ""]')
minrow = ws.min_row  # 最小行
maxrow = ws.max_row  # 最大行
mincol = ws.min_column  # 最小列
maxcol = ws.max_column  # 最大列
m = 3
n = 4
maxrow=int(maxrow/2)
for i in range(minrow, maxrow-1):
    # 录入姓名
    m += 2
    name = str(ws.cell(row=m, column=11).value)
    name ='"'+name+'"'
    #print(name)
    # 录入工号
    number = str(ws.cell(row=m, column=3).value)
    number = '"'+number+'"'
    # print('"'+number+'"')
    # 录入部门
    apartment = str(ws.cell(row=m, column=20).value)
    if(apartment=="None"):
        apartment='" "'
    else:
        apartment='"'+apartment+'"'
    # 录入上下班时间
    n += 2
    # 录入日期
    date = str(ws.cell(row=3, column=3).value)
    #print(date1[0:7])
    date2=date[0:4]
    # 数值保留两位
    date3=int(date[6:7])
    date3="{0:02d}".format(date3)
    #print(date3)
    date4=int(date[16:18])
    date4="{0:02d}".format(date4)
    date4=int(date4)
    #print(date4)
    for j in range(mincol, date4+1):
        date1 = ws.cell(row=4, column=j).value
        date = "{0:02d}".format(date1)
        date = date2+"-"+str(date3)+"-"+str(date)
        date = '"'+date+'"'
        #print(date)
        #print('"'+date+'"')
        #录入时间
        time = str(ws.cell(row=n, column=j).value)
        #print(len(time))
        if(len(time)==4):
            #print("空")
            #time='" "'+', " "'
            continue
        elif(len(time)==6):
            #print("上班时间："+time[0:5])
            time='"'+time[0:5]+'"'+', "'+time[0:5]+'"'
        elif(len(time)==12):
            time='"'+time[0:5]+'"'+', "'+time[6:11]+'"'
            #print("上班时间："+time[0:5])
            #print("下班时间："+time[6:11])
        elif(len(time)==18):
            time='"'+time[0:5]+'"'+', '+time[12:17]
            #print("上班时间："+time[0:5])
            #print("下班时间："+time[12:17])
        #print(time)
        #print("姓名："+name+",日期："+date+","+time+',备注：" "'+",工号："+number+",部门："+apartment)
        print(name+", "+date+", "+time+', " "'+", "+number+", "+apartment)
