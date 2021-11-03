import os
import json
import win32com.client as win32
path=os.getcwd()
path1="\\09月汇总表"
fname = path+path1
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
# FileFormat = 51 is for .xlsx extension
wb.SaveAs(fname, FileFormat = 51)
# FileFormat = 56 is for .xls extension
# wb.SaveAs(fname[:-1], FileFormat = 56)
wb.Close()
excel.Application.Quit()
print("转换成功！！！！")
# 删除原格式的表“09月汇总表.xls”
# os.remove("09月汇总表.xls")
from openpyxl import load_workbook
wb = load_workbook(filename='09月汇总表.xlsx')
ws = wb['刷卡记录']
print('记录格式：姓名，日期，上班时间，下班时间，备注，工号，部门')
print('[格式："戚厚之", "2021-09-07", "09:27", "18:54", "", "64", ""]')
minrow = ws.min_row  # 最小行
maxrow = ws.max_row  # 最大行
mincol = ws.min_column  # 最小列
maxcol = ws.max_column  # 最大列
m = 3
n = 4
maxrow=int(maxrow/2)
#新建一个表格
sheet=wb.active
sheet.title=("部门信息表")
with open('南京.json','r',encoding='utf-8') as f:
        #获取json内容
        data = json.load(f)
        #转为字符串
        #data = json.dumps(data,ensure_ascii=False)
        #转为dict
        #data = json.loads(data)
for i in range(minrow, maxrow-1):
    # 录入姓名 工号 部门
    m += 2
    name = str(ws.cell(row=m, column=11).value)
    def jobnumber(data,name):
        return [k for (k,v) in data.items() if v[0] == name]
    num=jobnumber(data,name)
    jobnumber = ','.join(num)
    # 录入工号
    #jobnumber = str(ws.cell(row=m, column=3).value)
    #jobnumber = '"'+jobnumber+'" '
    for key in data:
        #获取json中键值后面的值
        value=str(data[jobnumber])
    #print(value)
    #删除
    value1=value.replace('[','',1)
    value2=value1.replace(']','',1)
    value=value2.replace("'",'',4)
    #获取后半部分信息
    value1=value.split(", ")[1]
    department=value1
    #获取前半部分信息
    #value2=value.split(",")[0]
    #name=value2
    #result=jobnumber+", "+name+", "+apartment
    #content = result.split(',')
    # 录入部门
    #apartment = str(ws.cell(row=m, column=20).value)
    if(department=="None"):
        department=""
    else:
        department=department
    # 录入上下班时间
    n += 2
    # 录入日期
    date = str(ws.cell(row=3, column=3).value)
    date2=date[0:4]
    # 数值保留两位
    date3=int(date[6:7])
    date3="{0:02d}".format(date3)
    date4=int(date[16:18])
    date4="{0:02d}".format(date4)
    date4=int(date4)
    for j in range(mincol, date4+1):
        date1 = ws.cell(row=4, column=j).value
        date = "{0:02d}".format(date1)
        date = date2+"-"+str(date3)+"-"+str(date)
        #录入时间
        time = str(ws.cell(row=n, column=j).value)
        special=" "
        if(len(time)==4):
            continue
        elif(len(time)==6):
            sign_in=time[0:5]
            sign_out=time[0:5]
        elif(len(time)==12):
            sign_in=time[0:5]
            sign_out=time[6:11]
        elif(len(time)==18):
            sign_in=time[0:5]
            sign_out=time[12:17]
        #print("姓名："+name+",日期："+date+","+time+',备注：" "'+",工号："+jobnumber+",部门："+department)
        #print(name+", "+date+", "+sign_in+", "+sign_out+', " "'+", "+jobnumber+", "+department)
        import sqlite3
        #创建连接数据库
        conn = sqlite3.connect('attendance_sql.db')
        # 3、使用execute 方法执行 SQL 语句（DDL）
        sql_create= '''
            CREATE TABLE 'attendance_sheet' 
                        ('name' text,
                         'date' text,
                         'sign_in' text,
                         'sign_out' text,
                         'special' text,
                         'jobnumber' text,
                         'department' text)'''
        try:
        # 用 execute 执行一条 sql 语句
            conn.execute(sql_create)
        except:
            pass
        #print("创建成功！！！")
        # SQLite 可以忽略数据列的类型（即 name，pass，age）
        # '''insert语句 把一个新的行插入到表中'''
        sql_insert= ''' insert into attendance_sheet 
                      (name,
                       date,
                       sign_in,
                       sign_out,
                       special,
                       jobnumber,
                       department)
                  values
                      (:a_name, :a_date, :a_sign_in, :a_sign_out, :a_special, :a_jobnumber, :a_department)'''
        # 把数据保存
        conn.execute(sql_insert,{'a_name':name, 'a_date':date, 'a_sign_in':sign_in, 'a_sign_out':sign_out, 'a_special':special, 'a_jobnumber':jobnumber, 'a_department':department})
        #print("插入数据成功！！！")
        conn.commit()
        conn.close()
print("数据库连接成功！！！")
